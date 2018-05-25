#!/usr/bin/env python3.6
#

###################################
#     IMPORTACAO  DE LIBRARIES
###################################
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, GradientFill, Alignment

###################################
#  CRIA A DATA DO ARQ DE SAIDA
###################################
monthout = datetime.datetime.today().strftime('%y%m%d')
monthtvtout = datetime.datetime.today().strftime('%Y%m%d')

###################################
#  CRIA A DATA DO ARQ DE ENTRADA
###################################
yest=datetime.date.fromordinal(datetime.date.today().toordinal()-1)
yesterout=str(yest.strftime('%y%m%d'))

###################################
#        ESTILOS DE BORDA
###################################
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")
dashed = Side(border_style="dashed", color="000000")



###################################
# ESTILOS DE FORMATACAO DE CELULA
###################################

#### ESTILO 01 ####
mesfmt = NamedStyle(name = 'nm_mesfmt')
mesfmt.font = Font(name="Calibri", size=12)
mesfmt.border = Border(top=thin, left=thin, right=thin, bottom=thin)
mesfmt.alignment =  Alignment(horizontal="center", vertical="center", text_rotation=90)

#### ESTILO 02 ####
wkdfmt = NamedStyle(name = 'nm_wkdfmt')
wkdfmt.fill =  PatternFill("solid", fgColor="696969")
wkdfmt.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
wkdfmt.border = Border(top=thin, left=thin, right=thin, bottom=thin)
wkdfmt.alignment =  Alignment(horizontal="center", vertical="center", text_rotation=90)

#### ESTILO 03 ####
wekfmt = NamedStyle(name = 'nm_wekfmt')
wekfmt.fill =  PatternFill("solid", fgColor="808080")
wekfmt.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
wekfmt.alignment =  Alignment(horizontal="center", vertical="center")
wekfmt.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#### ESTILO 04 ####
hrendv = NamedStyle(name = 'nm_hrendv')
hrendv.fill =  PatternFill("solid", fgColor="C0C0C0")
hrendv.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#### ESTILO 05 ####
hrendok = NamedStyle(name = 'nm_hrendok')
hrendok.fill =  PatternFill("solid", fgColor="4682B4")
hrendok.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
hrendok.alignment =  Alignment(horizontal="center", vertical="center")
hrendok.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#### ESTILO 06 ####
hrendnok = NamedStyle(name = 'nm_hrendnok')
hrendnok.fill =  PatternFill("solid", fgColor="ADD8E6")
hrendnok.font = Font(name='Calibri', size=10, bold=True)
hrendnok.alignment =  Alignment(horizontal="center", vertical="center")
hrendnok.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#### ESTILO 07 ####
hrendout = NamedStyle(name = 'nm_hrendout')
hrendout.fill =  PatternFill("solid", fgColor="FF0000")
hrendout.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
hrendout.alignment =  Alignment(horizontal="center", vertical="center")
hrendout.border = Border(top=thin, left=thin, right=thin, bottom=thin)


###################################
# VARIAVEIS COM MAIOR FREQUENCIA
###################################
hoje = datetime.datetime.today().day

##################################
# VERIF QTDE DE LINHAS NA PLANILHA
#   E ARRUMA AS FALHAS DE BORDAS
###################################
conta = 0
row = 3

wb = load_workbook(filename = 'Bitacora'+yesterout+'.xlsx')
ws = wb.active
ws['A2'].border = Border(right=thin)
ws['C1'].border = Border(top=thin)
ws['D2'].border = Border(left=thin, right=thin )
ws['E2'].border = Border(left=thin, right=thin )

cola = "A"+str(row)
colc = "C"+str(row)

while ( ws[colc].value != None):
    ws[cola].border = Border(right=thin)

    row += 1
    colc = "C"+str(row)
    cola = "A"+str(row)
    conta +=1

conta += 2

ulti = conta +1
colb = "B"+str(ulti)
ws[colb].border = Border(top=thin)

leg = conta + 3
colc = "C"+str(leg)
ws[colc].border = Border(top=medium, bottom=medium )

cold = "D"+str(leg)
ws[cold].border = Border(top=medium, bottom=medium, right=medium )


print ("A PLANILHA TEM {} LINHAS.".format(conta))

###################################
#   DEFINE O PREFIXO DA CELULA
###################################
colpfx=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]

###################################
#  DEFINE CABEC DA COLUNA DIA/MES
###################################
row = 1
diam = datetime.datetime.today().strftime('%d/%b')

hoje = datetime.datetime.today().day
hojev = hoje + 5
coluna = '{}'.format(colpfx[hojev])
colrow = str(coluna)+str(row)

#print ("DIAM : {}".format(diam))
#print ("DIAM COLROW: {}".format(colrow))

ws[colrow] = diam
ws[colrow].style = mesfmt

###################################
#  DEFINE CABEC DA COLUNA SEMANA
###################################
row += 1
dweek=["Seg","Ter","Qua","Qui","Sex","Sab","Dom"]
diasmes = datetime.datetime.today().strftime('%Y,%m,%d')
ano, mes, dia  = diasmes.split(',')
dnumber=calendar.weekday(int(ano),int(mes),int(dia))

semana='{}'.format(dweek[dnumber])
#print ("SEMANA : {}".format(semana))

colrow = str(coluna)+str(row)
#print ("SEMANA COLROW: {}".format(colrow))

ws[colrow] = semana
if ( dnumber >= 5 ):
    ws[colrow].style = wkdfmt
else:
    ws[colrow].style = wekfmt

###################################
# CRIA A COLUNA FORMATADA E VAZIA
###################################
row = 3
hoje = datetime.datetime.today().day
hojev = hoje + 5
coluna = '{}'.format(colpfx[hojev])
colrow = str(coluna)+str(row)

#print ("HOJE : {}".format(hoje))
#print ("HOJEV : {}".format(hojev))
#print ("ROW : {}".format(row))
#print ("COLUNA : {}".format(coluna))
#print ("CONTA: {}".format(conta))
#print ("CONTA: {}".format(conta))

while ( row <= conta ):
    #print ("VAZIA : {}".format(colrow))
    ws[colrow].style = hrendv
    row += 1
    colrow = str(coluna)+str(row)

###################################
# VERIFICA A HORA FINAL NA PLANILHA
###################################
#print ("CONTA: {}".format(conta))
row = 3
jobcell = "C"+str(row)
metajob = "D"+str(row)
diajob = datetime.datetime.today().strftime('%Y-%m-%d')
#print ("DIAJOB: {}".format(diajob))

###################################
#  FAZ O APPEND DO ARQ INT E EXT
###################################
row = 3
ciclico = 0
jobcell = "C"+str(row)
hrinicio = "D"+str(row)
metajob = "F"+str(row)
diajob = datetime.datetime.today().strftime('%Y-%m-%d')
ontem = 0
interno = 0
print ("DIAJOB: {}".format(diajob))
while ( row <= conta ):
    if ( interno == 1 ):
        arq_csv = "ctm_bitacora_TIVIT_" + monthtvtout + ".csv"
    else:
        arq_csv = "ctm_bitacora_" + monthout + ".csv"
    for line in open(arq_csv):
        fields = line.strip().split(',')
        dtline = fields[1][0:10]
        if "CIP_NPC_SEND_VARREDURA_ADDA003" in ws[jobcell].value:
            #print("JOBFORA: {}".format(ws[jobcell].value))
            if ( ciclico == 0):
                jobname_old = ws[jobcell].value
                print("JOBOLD: {}".format(jobname_old))
            ciclico = 1
            ws[jobcell].value = "CIP_NPC_SEND_VARREDURA_ADDA003"
            #print ("######## CAIU AQUI !!!!! ########")
            #print ("HRINICIO: {}".format(ws[hrinicio].value))


        if ( dtline == diajob ) and ( ws[jobcell].value in line ) and (ws[hrinicio].value in line):
            #print ("JOBDENTRO: {}".format(ws[jobcell].value))
            #print ("======================= BEGIN =======================")
            #print ("LINEDENTRO: {}".format(line))
            start = fields[1]
            jobname = fields[0].strip().split(',')
            hrend = fields[2].strip().split(' ')
            horaf="{}".format(hrend[1][0:5])
            horafim = datetime.datetime.strptime(horaf, "%H:%M")
            #print("HORAFIM: {}".format(horafim))
            ###################################
            #    TRANSFORMA O CAMPO METAJOB
            ###################################
            metajobs = datetime.datetime.strptime(ws[metajob].value, "%H:%M:%S")
            #print("METAJOBS: {}".format(metajobs))
            ###################################
            # MUDA DE STR P TIME O ELAPSEDTIME
            ###################################
            elatime = str(fields[3].strip().split(','))
            elatime = elatime.replace('\'', '').replace(']', '').replace('[', '')
            #print("ELATIME: {}".format(elatime))
            elatime2 = datetime.datetime.strptime(elatime, "%H:%M:%S")
            #print("ELATIME2: {}".format(elatime2))
            ###################################
            #  VERIFICA SE E HOJE OU ONTEM
            ###################################
            if (ontem == 0):
                cold = hoje + 5
            else:
                cold = hoje + 4
            ##################################
            #  VERIF HORAFIM DENTRO DA META
            ##################################
            if ( elatime2 > metajobs ):
                coluna = '{}'.format(colpfx[cold])
                colrow = str(coluna)+str(row)
                ws[colrow] = horaf
                ws[colrow].style = hrendout
                #print("BRINCADEIRA ....")
                #print("HORAFIM: {}".format(horafim))
                #print("METAJOB: {}".format(ws[metajob].value))
                #print("FORA DA META - VERMELHO")
            else:
                ###################################
                # CRIA A VAR COM VLR FIXO DE 1 HRA
                ###################################
                umahora = "01:00:00"
                uhora = datetime.datetime.strptime(umahora, "%H:%M:%S")
                ###################################
                # VER JOBS NA META E ABXO DE 1 HRA
                ###################################
                if ( elatime2 > uhora ):
                    #print("EITAAAAA")
                    coluna = '{}'.format(colpfx[cold])
                    colrow = str(coluna)+str(row)
                    ws[colrow] = horaf
                    ws[colrow].style = hrendnok
                    #print ("DENTRO DA META MAIS QUE 1 HORA - AZUL CLARO")
                else:
                    #print("OH LOKO BICHO")
                    coluna = '{}'.format(colpfx[cold])
                    colrow = str(coluna)+str(row)
                    ws[colrow] = horaf
                    ws[colrow].style = hrendok
                    #print ("DENTRO DA META MENOS QUE 1 HORA - AZUL ESCUURO")
    row += 1
    if ( row == conta and ontem == 0 ):
        yest=datetime.date.fromordinal(datetime.date.today().toordinal()-1)
        diajob=str(yest.strftime('%Y-%m-%d'))
        row = 3
        ontem = 1
    if ( row == conta and ontem == 1 and interno == 0 ):
        diajob = datetime.datetime.today().strftime('%Y-%m-%d')
        row = 3
        ontem = 0
        interno = 1
    if ( ciclico == 1 ):
        ws[jobcell].value = jobname_old

    jobcell = "C"+str(row)
    hrinicio = "D"+str(row)
    metajob = "F"+str(row)
    ciclico = 0
#############################################
dest_filename = 'Bitacora'+monthout+'.xlsx'
wb.save(filename = dest_filename)
