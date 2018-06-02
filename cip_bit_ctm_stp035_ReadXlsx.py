#!/usr/bin/env python3.6
#

###################################
#     IMPORTACAO  DE LIBRARIES
###################################
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, GradientFill, Alignment

###################################
#  CRIA A DATA DO ARQ DE SAIDA
###################################
monthout = datetime.datetime.today().strftime('%y%m%d')
monthtvtout = datetime.datetime.today().strftime('%Y%m%d')
mthtday=datetime.datetime.today().strftime('%m')
###################################
#  CRIA A DATA DO ARQ DE ENTRADA
###################################
yest=datetime.date.fromordinal(datetime.date.today().toordinal()-1)
mthyest=str(yest.strftime('%m'))

if ( mthtday == mthyest ):
   yesterout=str(yest.strftime('%y%m%d'))
else:
   yesterout=monthout

###################################
#        ESTILOS DE BORDA
###################################
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")
dashed = Side(border_style="dashed", color="000000")

###################################
# VARIAVEIS COM MAIOR FREQUENCIA
###################################
hoje = datetime.datetime.today().day

##################################
# VERIF QTDE DE LINHAS NA PLANILHA
#   E ARRUMA AS FALHAS DE BORDAS
###################################
conta = 0
row = 3


wb = load_workbook(filename = 'Bitacora'+yesterout+'.xlsx')
ws = wb.active
ws['A2'].border = Border(right=thin)
ws['C1'].border = Border(top=thin)
ws['D2'].border = Border(left=thin, right=thin )
ws['E2'].border = Border(left=thin, right=thin )

cola = "A"+str(row)
colc = "C"+str(row)

while ( ws[colc].value != None):
    ws[cola].border = Border(right=thin)

    row += 1
    colc = "C"+str(row)
    cola = "A"+str(row)
    conta +=1

conta += 2

ulti = conta +1
colb = "B"+str(ulti)
ws[colb].border = Border(top=thin)

leg = conta + 3
colc = "C"+str(leg)
ws[colc].border = Border(top=medium, bottom=medium )

cold = "D"+str(leg)
ws[cold].border = Border(top=medium, bottom=medium, right=medium )


print ("A PLANILHA TEM {} LINHAS.".format(conta))

###################################
#   DEFINE O PREFIXO DA CELULA
###################################
colpfx=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]

###################################
#  DEFINE CABEC DA COLUNA DIA/MES
###################################
row = 1
diam = datetime.datetime.today().strftime('%d/%b')

hoje = datetime.datetime.today().day
hojev = hoje + 5
coluna = '{}'.format(colpfx[hojev])
colrow = str(coluna)+str(row)

#print ("DIAM : {}".format(diam))
#print ("DIAM COLROW: {}".format(colrow))

ws[colrow] = diam
ws[colrow].font = Font(name="Calibri", size=12)
ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
ws[colrow].alignment =  Alignment(horizontal="center", vertical="center", text_rotation=90)

###################################
#  DEFINE CABEC DA COLUNA SEMANA
###################################
row += 1
dweek=["Seg","Ter","Qua","Qui","Sex","Sab","Dom"]
diasmes = datetime.datetime.today().strftime('%Y,%m,%d')
ano, mes, dia  = diasmes.split(',')
dnumber=calendar.weekday(int(ano),int(mes),int(dia))

semana='{}'.format(dweek[dnumber])
#print ("SEMANA : {}".format(semana))

colrow = str(coluna)+str(row)
#print ("SEMANA COLROW: {}".format(colrow))

ws[colrow] = semana
if ( dnumber >= 5 ):
    ws[colrow].fill =  PatternFill("solid", fgColor="696969")
    ws[colrow].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[colrow].alignment =  Alignment(horizontal="center", vertical="center", text_rotation=90)
else:
    ws[colrow].fill =  PatternFill("solid", fgColor="808080")
    ws[colrow].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[colrow].alignment =  Alignment(horizontal="center", vertical="center")
    ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)

###################################
# CRIA A COLUNA FORMATADA E VAZIA
###################################
row = 3
hoje = datetime.datetime.today().day
hojev = hoje + 5
coluna = '{}'.format(colpfx[hojev])
colrow = str(coluna)+str(row)

#print ("HOJE : {}".format(hoje))
#print ("HOJEV : {}".format(hojev))
#print ("ROW : {}".format(row))
#print ("COLUNA : {}".format(coluna))
#print ("CONTA: {}".format(conta))
#print ("CONTA: {}".format(conta))

while ( row <= conta ):
    #print ("VAZIA : {}".format(colrow))
    ws[colrow].fill =  PatternFill("solid", fgColor="C0C0C0")
    ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row += 1
    colrow = str(coluna)+str(row)

###################################
# VERIFICA A HORA FINAL NA PLANILHA
###################################
#print ("CONTA: {}".format(conta))
row = 3
jobcell = "C"+str(row)
metajob = "D"+str(row)
#diajob = datetime.datetime.today().strftime('%Y-%m-%d')
#print ("DIAJOB: {}".format(diajob))

###################################
#  FAZ O APPEND DO ARQ INT E EXT
###################################
row = 3
ciclico = 0
jobcell = "C"+str(row)
hrinicio = "D"+str(row)
metajob = "F"+str(row)
diajob = datetime.datetime.today().strftime('%Y-%m-%d')
ontem = 0
interno = 0
print ("DIAJOB: {}".format(diajob))
while ( row <= conta ):
    if ( interno == 1 ):
        arq_csv = "ctm_bitacora_TIVIT_" + monthtvtout + ".csv"
    else:
        arq_csv = "ctm_bitacora_" + monthout + ".csv"
    for line in open(arq_csv):
        fields = line.strip().split(',')
        dtline = fields[1][0:10]
        mline = fields[1][6:7]
        if (int(mline) < 10):
            mthline = "0" + str(mline)
        if "CIP_NPC_SEND_VARREDURA_ADDA003" in ws[jobcell].value:
            if ( ciclico == 0):
                jobname_old = ws[jobcell].value
            ciclico = 1
            ws[jobcell].value = "CIP_NPC_SEND_VARREDURA_ADDA003"
        #print ("DTLINE:  {} DIAJOB: {} JOBCELL: {} HRINICIO: {}".format(dtline, diajob, ws[jobcell].value, ws[hrinicio].value))
        if ( dtline == diajob ) and ( ws[jobcell].value in line ) and (ws[hrinicio].value in line):
            print ("DTLINE {} DIAJOB {}".format(dtline, diajob))
            print ("LINEDENTRO: {}".format(line))
            start = fields[1]
            jobname = fields[0].strip().split(',')
            hrend = fields[2].strip().split(' ')
            horaf="{}".format(hrend[1][0:5])
            horafim = datetime.datetime.strptime(horaf, "%H:%M")
            metajobs = datetime.datetime.strptime(ws[metajob].value, "%H:%M:%S")
            elatime = str(fields[3].strip().split(','))
            elatime = elatime.replace('\'', '').replace(']', '').replace('[', '')
            elatime2 = datetime.datetime.strptime(elatime, "%H:%M:%S")
            ###################################
            #  VERIFICA SE E HOJE OU ONTEM
            ###################################
            if (ontem == 0):
                #print ("HOJE  DIAJOB {} ROW1 {} CONTA {} ONTEM {} INTERNO {}".format(diajob, row, conta, ontem, interno))
                cold = hoje + 5
            else:
                #print ("ONTEM DIAJOB {} ROW1 {} CONTA {} ONTEM {} INTERNO {}".format(diajob, row, conta, ontem, interno))
                yestday=str(yest.strftime('%d'))
                print ("YESTDAY {}".format(yestday))
                cold = int(yestday) + 5
            ##################################
            #  VERIF HORAFIM DENTRO DA META
            ##################################
            if ( elatime2 > metajobs ):
                coluna = '{}'.format(colpfx[cold])
                colrow = str(coluna)+str(row)
                print ("COLROW1...: {}".format(colrow))
                ws[colrow] = horaf
                ws[colrow].fill =  PatternFill("solid", fgColor="FF0000")
                ws[colrow].font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                ws[colrow].alignment =  Alignment(horizontal="center", vertical="center")
                ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                ###################################
                # CRIA A VAR COM VLR FIXO DE 1 HRA
                ###################################
                umahora = "01:00:00"
                uhora = datetime.datetime.strptime(umahora, "%H:%M:%S")
                ###################################
                # VER JOBS NA META E ABXO DE 1 HRA
                ###################################
                if ( elatime2 > uhora ):
                    #print("EITAAAAA")
                    coluna = '{}'.format(colpfx[cold])
                    colrow = str(coluna)+str(row)
                    print ("COLROW2...: {}".format(colrow))
                    ws[colrow] = horaf
                    ws[colrow].fill =  PatternFill("solid", fgColor="ADD8E6")
                    ws[colrow].font = Font(name='Calibri', size=10, bold=True)
                    ws[colrow].alignment =  Alignment(horizontal="center", vertical="center")
                    ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                else:
                    coluna = '{}'.format(colpfx[cold])
                    colrow = str(coluna)+str(row)
                    print ("COLROW3...: {}".format(colrow))
                    ws[colrow] = horaf
                    ws[colrow].fill =  PatternFill("solid", fgColor="4682B4")
                    ws[colrow].font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                    ws[colrow].alignment =  Alignment(horizontal="center", vertical="center")
                    ws[colrow].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ##############################################
            #   VERIFICA SE EXISTE A QUEBRA DO MES
            ##############################################
            mthtday = datetime.datetime.today().strftime('%m')
            #tday = datetime.date.fromordinal(datetime.date.today(%m))

            if ( mthline == mthtday ):
                #print ("YEST {} e TDAY {}: ".format (yest,tday))
                print("SEM QUEBRA DE MES ==>  MTHDAY {} MTHLINE {}".format(mthtday,mthline))
                monthout = datetime.datetime.today().strftime('%y%m%d')
                dest_filename = 'Bitacora'+monthout+'.xlsx'
                wb.save(filename = dest_filename)
                print ("Linha {} Arquivo {} gravado com sucesso.".format(row, dest_filename))
                print("###################################################################")
            else:
                print("COM QUEBRA DE MES ==>  MTHDAY {} MTHLINE {}".format(mthtday,mthline))
                monthyest=str(yest.strftime('%y%m%d'))
                print ("MONTHEND : {}".format(monthyest))
                dest_filename = 'Bitacora'+monthyest+'.xlsx'
                wb.save(filename = dest_filename)
                print ("Linha {} Arquivo {} gravado com sucesso.".format(row, dest_filename))
                print("###################################################################")
    row += 1
    if ( row == conta and ontem == 0 and interno == 0 ):
        wb.close()
        yest=datetime.date.fromordinal(datetime.date.today().toordinal()-1)
        diajob=str(yest.strftime('%Y-%m-%d'))
        yesterout = str(yest.strftime('%y%m%d'))
        wb = load_workbook(filename='Bitacora' + yesterout + '.xlsx')
        ws = wb.active
        ontem = 1
        row = 3

    if ( row == conta and ontem == 1 and interno == 0 ):
        wb.close()
        diajob = datetime.datetime.today().strftime('%Y-%m-%d')
        todayout = datetime.datetime.today().strftime('%y%m%d')
        wb = load_workbook(filename='Bitacora' + todayout + '.xlsx')
        ws = wb.active
        ontem = 0
        interno = 1
        row = 3

    if ( row == conta and ontem == 0 and interno == 1 ):
        wb.close()
        yest=datetime.date.fromordinal(datetime.date.today().toordinal()-1)
        diajob=str(yest.strftime('%Y-%m-%d'))
        yesterout = str(yest.strftime('%y%m%d'))
        wb = load_workbook(filename='Bitacora' + yesterout + '.xlsx')
        ws = wb.active
        ontem = 1
        row = 3


################################################
################################################
################################################

    if ( ciclico == 1 ):
        ws[jobcell].value = jobname_old

    jobcell = "C"+str(row)
    hrinicio = "D"+str(row)
    metajob = "F"+str(row)
    ciclico = 0
#############################################
